## python 2.7.15
from openpyxl import load_workbook
import pyperclip

## bros is a spreadsheet that represents all active members of the fraternity
## sizes is a spreadsheet that shows who has provided their t shirt sizes.

## so if someone is in bros, but not in sizes, they haven't provided their
## shirt size to me
 
bros = load_workbook('bros.xlsx').active
sizes = load_workbook('sizes.xlsx').active

## Those who have not responded to my email will go into this array
not_filled = []

for active in bros.rows:
    ## If flag is still true at the end of a loop, then that brother never
    ## responded to my email.
    flag = True
    for shirt in sizes.rows:
        ## All values are put in lowercase because names aren't all in the same
        ## case in the spreadsheets.
        aval = active[0].value.lower()
        sval = shirt[0].value.lower()

        ## this will find the email of the brother
        aat = aval.find('@')
        sat = sval.find('@')

        ## emails are truncated to just the name
        aval = aval[0:aat]
        sval = sval[0:sat]

        if aval == sval:
            flag = False
            break
    if flag:
        not_filled.append(active[0].value)

## this will be our string formatted for me to copy and batch email
out = ''

for fil in not_filled:
    out += fil + ','

## I don't want a comma at the end of the list of brothers, so here's some
## readability.
no_comma = len(out) - 1

out = out[0:no_comma]

## the list of brothers that haven't responded as an array.
def get_list():
    return not_filled

## names of brothers that have not responded.
def get_names():
    return out

## number of brothers that haven't responded.
def get_count():
    return len(not_filled)


if __name__ == "__main__":
    print get_count()
    print get_names()
    ## raw input is so the script doesn't just quit out in my terminal
    raw_input()
